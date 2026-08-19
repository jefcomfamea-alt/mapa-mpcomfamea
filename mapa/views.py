from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import logout
from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.db.models import Q
from django.http import HttpResponseForbidden
from datetime import timedelta
from django.utils import timezone
from urllib.parse import urlencode

from .forms import (
    CasoForm,
    UbicacionPreliminarForm
)
from .models import (
    Caso,
    SolicitudModificacion,
    Mensaje,
    Region,
    Provincia,
    Distrito,
    UbicacionPreliminar,
    PersonaPreliminar,
    RolPersonaPreliminar,
    Agresor
)
from .decorators import grupo_requerido
from .forms_solicitudes import SolicitudModificacionForm

def cargar_provincias(request):
    region_id = request.GET.get("region_id")

    provincias = Provincia.objects.filter(
        region_id=region_id
    ).order_by("nombre")

    data = [
        {
            "id": provincia.id,
            "nombre": provincia.nombre
        }
        for provincia in provincias
    ]

    return JsonResponse(data, safe=False)

def cargar_distritos(request):
    provincia_id = request.GET.get("provincia_id")

    distritos = Distrito.objects.filter(
        provincia_id=provincia_id
    ).order_by("nombre")

    data = [
        {
            "id": distrito.id,
            "nombre": distrito.nombre
        }
        for distrito in distritos
    ]

    return JsonResponse(data, safe=False)

@login_required
def inicio(request):

    pendientes_mensajes = Mensaje.objects.filter(
        destinatario=request.user,
        leido=False
    ).count()

    hoy = timezone.now().date()
    limite = hoy + timedelta(days=3)

    if (
        request.user.is_superuser
        or request.user.groups.filter(
            name__in=["Administrador", "Jefe_MP"]
        ).exists()
    ):

        notificaciones = Caso.objects.filter(
            estado="ACTIVO",
            fecha_limite__isnull=False,
            fecha_limite__lte=limite
        ).filter(
            Q(fecha_registro__year__gte=2026) |
            Q(ultima_visita__year__gte=2026)
        ).order_by("fecha_limite")

    else:

        notificaciones = Caso.objects.filter(
            estado="ACTIVO",
            responsable=request.user,
            fecha_limite__isnull=False,
            fecha_limite__lte=limite
        ).filter(
            Q(fecha_registro__year__gte=2026) |
            Q(ultima_visita__year__gte=2026)
        ).order_by("fecha_limite")


    pendientes_casos = notificaciones.count()

    pendientes = pendientes_mensajes + pendientes_casos


    es_admin = (
        request.user.is_superuser
        or request.user.groups.filter(
            name="Administrador"
        ).exists()
    )

    es_jefe = request.user.groups.filter(
        name="Jefe_MP"
    ).exists()

    es_usuario_mp = request.user.groups.filter(
        name="Usuario_MP"
    ).exists()

    es_efectivo_comfamea = request.user.groups.filter(
        name="Efectivo_COMFAMEA"
    ).exists()

    es_usuario_investigacion = request.user.groups.filter(
        name="Usuario_Investigacion"
    ).exists()

    es_usuario_estadistica = request.user.groups.filter(
        name="Usuario_Estadistica"
    ).exists()

    return render(
        request,
        "mapa/inicio.html",
        {
            "pendientes": pendientes,
            "es_admin": es_admin,
            "es_jefe": es_jefe,
            "es_usuario_mp": es_usuario_mp,
            "es_efectivo_comfamea": es_efectivo_comfamea,
            "es_usuario_investigacion": es_usuario_investigacion,
            "es_usuario_estadistica": es_usuario_estadistica,
            "notificaciones": notificaciones,
        }
    )

@login_required
def notificaciones(request):

    hoy = timezone.now().date()
    limite = hoy + timedelta(days=3)

    if (
        request.user.is_superuser
        or request.user.groups.filter(
            name__in=["Administrador", "Jefe_MP"]
        ).exists()
    ):

        notificaciones = Caso.objects.filter(
            estado="ACTIVO",
            fecha_limite__isnull=False,
            fecha_limite__lte=limite
        ).filter(
            Q(fecha_registro__year__gte=2026) |
            Q(ultima_visita__year__gte=2026)
        ).order_by("fecha_limite")

    else:

        notificaciones = Caso.objects.filter(
            estado="ACTIVO",
            responsable=request.user,
            fecha_limite__isnull=False,
            fecha_limite__lte=limite
        ).filter(
            Q(fecha_registro__year__gte=2026) |
            Q(ultima_visita__year__gte=2026)
        ).order_by("fecha_limite")

    return render(
        request,
        "mapa/notificaciones.html",
        {
            "notificaciones": notificaciones,
            "hoy": hoy,
        }
    )

@login_required
@grupo_requerido("Administrador")
def administrar_usuarios(request):

    usuarios = User.objects.all().order_by("username")

    grupos = Group.objects.all().order_by("name")

    return render(
        request,
        "mapa/administrar_usuarios.html",
        {
            "usuarios": usuarios,
            "grupos": grupos
        }
    )

@login_required
@grupo_requerido("Administrador")
def editar_usuario(request, pk):

    usuario = get_object_or_404(User, pk=pk)

    if request.method == "POST":

        usuario.username = request.POST.get("username")
        usuario.first_name = request.POST.get("first_name")
        usuario.last_name = request.POST.get("last_name")
        usuario.email = request.POST.get("email")

        password = request.POST.get("password")

        if password:
            usuario.set_password(password)

        usuario.save()

        rol = request.POST.get("rol")

        if rol:

            grupo = Group.objects.get(name=rol)

            usuario.groups.clear()

            usuario.groups.add(grupo)

        return redirect("administrar_usuarios")

    grupo_actual = usuario.groups.first()

    return render(
        request,
        "mapa/editar_usuario.html",
        {
            "usuario": usuario,
            "grupo_actual": grupo_actual,
        }
    )

@login_required
@grupo_requerido("Administrador")
def nuevo_usuario(request):

    if request.method == "POST":

        username = request.POST.get("username")
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        rol = request.POST.get("rol")

        usuario = User.objects.create_user(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            password=password
        )

        grupo = Group.objects.get(name=rol)

        usuario.groups.add(grupo)

        return redirect("administrar_usuarios")

    grupos = Group.objects.all().order_by("name")

    return render(
        request,
        "mapa/nuevo_usuario.html",
        {
            "grupos": grupos
        }
    )

@login_required
@grupo_requerido("Administrador")
def desactivar_usuario(request, pk):

    usuario = get_object_or_404(User, pk=pk)

    usuario.is_active = False
    usuario.save()

    return redirect("administrar_usuarios")

@login_required
@grupo_requerido("Administrador", "Jefe_MP", "Usuario_MP")
def nuevo_caso(request):

    # ==========================================
    # DATOS RECIBIDOS DESDE CONVERSIÓN
    # ==========================================

    preliminar_id = request.GET.get("preliminar")
    rol = request.GET.get("rol")
    personas_param = request.GET.get("personas")

    preliminar = None
    beneficiario = None
    agresor = None

    # ==========================================
    # INTERPRETAR ROL
    # ==========================================

    def interpretar_rol(tipo):

        tipo = (tipo or "").strip().upper()

        tipo = (
            tipo.replace("/", "")
                .replace("-", "")
                .replace("_", "")
                .replace(" ", "")
        )

        if tipo in [
            "PRESUNTAVICTIMA",
            "BENEFICIARIA",
            "BENEFICIARIO"
        ]:
            return "BENEFICIARIO"

        if tipo in [
            "DENUNCIADO",
            "AGRESORA",
            "AGRESOR"
        ]:
            return "AGRESOR"

        if tipo in [
            "AMBOS",
            "AMBAS"
        ]:
            return "AMBOS"

        return None

    # ==========================================
    # CARGAR PERSONA PRELIMINAR
    # ==========================================

    if preliminar_id:

        preliminar = get_object_or_404(
            PersonaPreliminar,
            pk=preliminar_id
        )

        rol_preliminar = interpretar_rol(rol)

        if rol_preliminar == "BENEFICIARIO":

            beneficiario = preliminar

        elif rol_preliminar == "AGRESOR":

            agresor = preliminar

        elif rol_preliminar == "AMBOS":

            beneficiario = preliminar
            agresor = preliminar

    # ==========================================
    # CARGAR PERSONAS RELACIONADAS
    # ==========================================

    elif personas_param:

        for dato in personas_param.split(","):

            try:

                persona_id, tipo = dato.split(":", 1)

                persona = PersonaPreliminar.objects.get(
                    pk=int(persona_id)
                )

                rol_persona = interpretar_rol(tipo)

                if rol_persona == "BENEFICIARIO":

                    beneficiario = persona

                elif rol_persona == "AGRESOR":

                    agresor = persona

                elif rol_persona == "AMBOS":

                    beneficiario = persona
                    agresor = persona

            except Exception as e:

                print(
                    "ERROR INTERPRETANDO PERSONA:",
                    dato,
                    e
                )

    # ==========================================
    # GUARDAR CASO
    # ==========================================

    if request.method == "POST":

        form = CasoForm(
            request.POST,
            usuario=request.user
        )

        if form.is_valid():

            caso = form.save(commit=False)

            # ==========================================
            # COORDENADAS DE PERSONA PRELIMINAR
            # ==========================================

            if agresor:

                caso.latitud_agresor = agresor.latitud
                caso.longitud_agresor = agresor.longitud

                print(
                    "COORDENADAS AGRESOR PRELIMINAR:",
                    agresor.nombres,
                    agresor.latitud,
                    agresor.longitud
                )

            if beneficiario:

                caso.latitud = beneficiario.latitud
                caso.longitud = beneficiario.longitud

                print(
                    "COORDENADAS BENEFICIARIO PRELIMINAR:",
                    beneficiario.nombres,
                    beneficiario.latitud,
                    beneficiario.longitud
                )

            # ==========================================
            # RESPONSABLE
            # ==========================================

            if not caso.responsable:
                caso.responsable = request.user

            if caso.responsable:

                nombre = (
                    f"{caso.responsable.first_name} "
                    f"{caso.responsable.last_name}"
                ).strip()

                caso.efectivo = (
                    nombre
                    if nombre
                    else caso.responsable.username
                )

                caso.edicion_autorizada = False

            caso.save()

            form.save_m2m()

            # ==========================================
            # REGISTRAR AGRESOR DEFINITIVO
            # ==========================================

            if caso.agresor:

                agresor_obj, creado = Agresor.objects.get_or_create(

                    dni=caso.dni_agresor,

                    defaults={
                        "nombres": caso.agresor,
                        "domicilio": caso.direccion_agresor,
                        "latitud": caso.latitud_agresor,
                        "longitud": caso.longitud_agresor,
                        "activo": True,
                    }
                )

                if not creado:

                    agresor_obj.nombres = caso.agresor
                    agresor_obj.domicilio = caso.direccion_agresor
                    agresor_obj.latitud = caso.latitud_agresor
                    agresor_obj.longitud = caso.longitud_agresor
                    agresor_obj.activo = True

                    agresor_obj.save()

                caso.agresor_registro = agresor_obj

                caso.save(
                    update_fields=["agresor_registro"]
                )

            # ==========================================
            # MARCAR PERSONAS REALMENTE CONVERTIDAS
            # ==========================================

            personas_convertidas = []

            if personas_param:

                for dato in personas_param.split(","):

                    try:

                        persona_id, tipo = dato.split(":", 1)

                        persona = PersonaPreliminar.objects.get(
                            pk=int(persona_id)
                        )

                        persona.convertida = True

                        persona.save(
                            update_fields=["convertida"]
                        )

                        personas_convertidas.append(persona)

                    except Exception as e:

                        print(
                            "ERROR MARCANDO PERSONA CONVERTIDA:",
                            dato,
                            e
                        )

            elif preliminar:

                preliminar.convertida = True

                preliminar.save(
                    update_fields=["convertida"]
                )

                personas_convertidas.append(preliminar)

            # ==========================================
            # ACTUALIZAR UBICACIÓN PRELIMINAR
            # ==========================================

            ubicacion = None

            if personas_param:

                primer_dato = personas_param.split(",")[0]

                try:

                    persona_id, tipo = primer_dato.split(":", 1)

                    persona_convertida = PersonaPreliminar.objects.get(
                        pk=int(persona_id)
                    )

                    ubicacion = persona_convertida.ubicacion_preliminar

                except Exception as e:

                    print(
                        "ERROR OBTENIENDO UBICACION:",
                        e
                    )

            elif preliminar:

                ubicacion = preliminar.ubicacion_preliminar   


            if ubicacion:

                pendientes = PersonaPreliminar.objects.filter(
                    ubicacion_preliminar=ubicacion,
                    convertida=False
                ).exists()

                if not pendientes:

                    ubicacion.estado = "CONVERTIDA"
                    ubicacion.caso_generado = caso
                    ubicacion.convertida_por = request.user
                    ubicacion.fecha_conversion = timezone.now()

                    ubicacion.save()

            # ==========================================
            # MENSAJE AL RESPONSABLE
            # ==========================================

            if caso.responsable:

                Mensaje.objects.create(

                    destinatario=caso.responsable,

                    caso=caso,

                    asunto="🔔 Nuevo caso asignado",

                    contenido=(
                        f"Se le comunica que se le ha asignado "
                        f"el expediente N.° "
                        f"{caso.expediente or caso.folder} "
                        f"para realizar la ejecución y seguimiento "
                        f"de las medidas de protección."
                    )
                )

            return redirect("inicio")

        else:

            print(form.errors)

    # ==========================================
    # MOSTRAR FORMULARIO
    # ==========================================

    else:

        lat = request.GET.get("lat")
        lng = request.GET.get("lng")

        datos_iniciales = {
            "latitud": lat,
            "longitud": lng,
        }

        # ==========================================
        # BENEFICIARIA
        # ==========================================

        if beneficiario:

            datos_iniciales.update({

                "beneficiario": beneficiario.nombres,
                "dni_beneficiario": beneficiario.dni,
                "telefono": beneficiario.telefono,
                "domicilio": beneficiario.direccion,

                "latitud": (
                    beneficiario.latitud
                    if beneficiario.latitud is not None
                    else lat
                ),

                "longitud": (
                    beneficiario.longitud
                    if beneficiario.longitud is not None
                    else lng
                ),
            })

        # ==========================================
        # AGRESOR
        # ==========================================

        if agresor:

            datos_iniciales.update({

                "agresor": agresor.nombres,
                "dni_agresor": agresor.dni,
                "telefono_agresor": agresor.telefono,
                "direccion_agresor": agresor.direccion,

                "latitud_agresor": agresor.latitud,
                "longitud_agresor": agresor.longitud,
            })

        # ==========================================
        # PERSONA PRELIMINAR
        # ==========================================

        if preliminar:

            rol_preliminar = interpretar_rol(rol)

            if rol_preliminar == "BENEFICIARIO":

                datos_iniciales.update({

                    "beneficiario": preliminar.nombres,
                    "dni_beneficiario": preliminar.dni,
                    "telefono": preliminar.telefono,
                    "domicilio": preliminar.direccion,
                })

            elif rol_preliminar == "AGRESOR":

                datos_iniciales.update({

                    "agresor": preliminar.nombres,
                    "dni_agresor": preliminar.dni,
                    "telefono_agresor": preliminar.telefono,
                    "direccion_agresor": preliminar.direccion,
                })

            elif rol_preliminar == "AMBOS":

                datos_iniciales.update({

                    "beneficiario": preliminar.nombres,
                    "dni_beneficiario": preliminar.dni,
                    "telefono": preliminar.telefono,
                    "domicilio": preliminar.direccion,

                    "agresor": preliminar.nombres,
                    "dni_agresor": preliminar.dni,
                    "telefono_agresor": preliminar.telefono,
                    "direccion_agresor": preliminar.direccion,
                })

            if preliminar.latitud is not None:

                datos_iniciales["latitud"] = (
                    preliminar.latitud
                )

            if preliminar.longitud is not None:

                datos_iniciales["longitud"] = (
                    preliminar.longitud
                )

        # ==========================================
        # FORMULARIO
        # ==========================================

        form = CasoForm(
            usuario=request.user,
            initial=datos_iniciales
        )

    return render(
        request,
        "mapa/nuevo_caso.html",
        {
            "form": form,
            "preliminar": preliminar,
            "beneficiario": beneficiario,
            "agresor": agresor,
            "rol": rol,
        }
    )

@login_required
@grupo_requerido(
    "Administrador",
    "Jefe_MP",
    "Usuario_Investigacion"
)
def registrar_ubicacion_preliminar(request):

    UbicacionPreliminar.objects.filter(
        estado="PRELIMINAR",
        fecha_vencimiento__lt=timezone.now().date()
    ).delete()


    if request.method == "POST":

        form = UbicacionPreliminarForm(request.POST)

        if form.is_valid():

            ubicacion = form.save(commit=False)

            ubicacion.registrado_por = request.user

            ubicacion.latitud = form.cleaned_data["latitud"]
            ubicacion.longitud = form.cleaned_data["longitud"]

            ubicacion.save()


            # ==============================
            # GUARDAR PERSONAS INVOLUCRADAS
            # ==============================

            nombres = request.POST.getlist("nombres")
            dni = request.POST.getlist("dni")
            telefono = request.POST.getlist("telefono")
            direccion = request.POST.getlist("direccion")
            roles = request.POST.getlist("rol")
            latitudes = request.POST.getlist("latitud_persona")
            longitudes = request.POST.getlist("longitud_persona")


            for i in range(len(nombres)):

                if nombres[i]:

                    persona = PersonaPreliminar.objects.create(

                        ubicacion_preliminar=ubicacion,

                        nombres=nombres[i],

                        dni=dni[i] if i < len(dni) else "",

                        telefono=telefono[i] if i < len(telefono) else "",

                        direccion=direccion[i] if i < len(direccion) else "",

                        latitud=(
                            latitudes[i]
                            if i < len(latitudes) and latitudes[i]
                            else None
                        ),

                        longitud=(
                            longitudes[i]
                            if i < len(longitudes) and longitudes[i]
                            else None
                        )

                    )


                    if i < len(roles) and roles[i]:

                        RolPersonaPreliminar.objects.create(

                            persona=persona,

                            rol=roles[i]

                        )


            return redirect("inicio")


    else:

        form = UbicacionPreliminarForm()


    return render(
        request,
        "mapa/registrar_ubicacion_preliminar.html",
        {
            "form": form
        }
    )

@login_required
@grupo_requerido("Administrador", "Jefe_MP")
def seleccionar_rol_preliminar(request, id):

    persona = get_object_or_404(
        PersonaPreliminar,
        pk=id
    )

    # Todas las personas de la misma ubicación preliminar
    personas = PersonaPreliminar.objects.filter(
        ubicacion_preliminar=persona.ubicacion_preliminar,
        convertida=False
    ).prefetch_related("roles").order_by("id")


    # =====================================================
    # PROCESAR LOS ROLES SELECCIONADOS
    # =====================================================

    if request.method == "POST":

        personas_param = []

        for p in personas:

            rol = request.POST.get(
                f"rol_{p.id}"
            )

            if rol:

                personas_param.append(
                    f"{p.id}:{rol}"
                )


        # ==============================================
        # VERIFICAR QUE SE HAYA SELECCIONADO UN ROL
        # ==============================================

        if not personas_param:

            return render(
                request,
                "mapa/seleccionar_rol_preliminar.html",
                {
                    "persona": persona,
                    "personas": personas,
                    "error": "Debe seleccionar el rol de al menos una persona.",
                }
            )


        # ==============================================
        # ENVIAR LAS PERSONAS A NUEVO CASO
        # ==============================================

        parametros = urlencode({
            "personas": ",".join(personas_param)
        })

        return redirect(
            f"/nuevo-caso/?{parametros}"
        )


    # =====================================================
    # MOSTRAR PÁGINA
    # =====================================================

    return render(
        request,
        "mapa/seleccionar_rol_preliminar.html",
        {
            "persona": persona,
            "personas": personas,
        }
    )

@login_required
@grupo_requerido(
    "Administrador",
    "Jefe_MP",
    "Usuario_MP",
    "Usuario_Estadistica"
)
def gestion_casos(request):

    busqueda = request.GET.get("q", "").strip()

    casos = Caso.objects.all()

    if busqueda:

        casos = casos.filter(
            Q(beneficiario__icontains=busqueda) |
            Q(dni_beneficiario__icontains=busqueda) |
            Q(agresor__icontains=busqueda) |
            Q(dni_agresor__icontains=busqueda) |
            Q(expediente__icontains=busqueda) |
            Q(folder__icontains=busqueda)
        )

    casos = casos.order_by("-id")

    # ============================
    # PAGINACIÓN: 20 CASOS
    # ============================
    from django.core.paginator import Paginator

    paginator = Paginator(casos, 20)

    pagina = request.GET.get("page")

    casos = paginator.get_page(pagina)

    return render(
        request,
        "mapa/gestion_casos.html",
        {
            "casos": casos,
            "busqueda": busqueda,
        }
    )

@login_required
@grupo_requerido("Administrador", "Jefe_MP")
def casos_preliminares(request):

    ubicaciones = (
        UbicacionPreliminar.objects
        .filter(
            estado="PRELIMINAR"
        )
        .prefetch_related(
            "personas__roles"
        )
        .order_by("-fecha_registro", "-id")
    )

    # Solo mostramos ubicaciones que todavía tengan
    # personas preliminares sin convertir
    preliminares = []

    for ubicacion in ubicaciones:

        personas = ubicacion.personas.filter(
            convertida=False
        ).prefetch_related("roles")

        if personas.exists():

            preliminares.append({
                "ubicacion": ubicacion,
                "personas": personas,
            })

    # ==========================================
    # PAGINACIÓN: 20 CASOS PRELIMINARES POR PÁGINA
    # ==========================================

    from django.core.paginator import Paginator

    paginator = Paginator(preliminares, 20)

    pagina = request.GET.get("page")

    preliminares = paginator.get_page(pagina)

    return render(
        request,
        "mapa/casos_preliminares.html",
        {
            "preliminares": preliminares,
        }
    )

@login_required
@grupo_requerido("Administrador")
def eliminar_caso_preliminar(request, id):

    ubicacion = get_object_or_404(
        UbicacionPreliminar,
        pk=id,
        estado="PRELIMINAR"
    )

    if request.method == "POST":

        # ==========================================
        # ELIMINAR PERSONAS ASOCIADAS
        # ==========================================

        personas = PersonaPreliminar.objects.filter(
            ubicacion_preliminar=ubicacion
        )

        personas.delete()

        # ==========================================
        # ELIMINAR UBICACIÓN PRELIMINAR
        # ==========================================

        ubicacion.delete()

        return redirect("casos_preliminares")

    return redirect("casos_preliminares")


@login_required
def casos_json(request):

    features = []

    for caso in Caso.objects.filter(estado="ACTIVO"):

        # ==========================================
        # 🔵 UBICACIÓN DE LA BENEFICIARIA
        # ==========================================

        if caso.latitud is not None and caso.longitud is not None:

            features.append({
                "type": "Feature",

                "geometry": {
                    "type": "Point",
                    "coordinates": [
                        caso.longitud,
                        caso.latitud
                    ]
                },

                "properties": {
                    "id": caso.id,
                    "TIPO_PUNTO": "BENEFICIARIO",

                    "BENEFICIARIO": caso.beneficiario,
                    "DOMICILIO": caso.domicilio,
                    "NIVEL RIESGO": caso.nivel_riesgo,
                    "DISTRITO": str(caso.distrito_denuncia)
                        if caso.distrito_denuncia else "",

                    "COMISARIA DENUNCIA": caso.comisaria_denuncia,
                    "COMISARIA MEDIDA": caso.comisaria_medida,

                    "EFECTIVO": caso.efectivo,
                    "RESPONSABLE_ID":
                        caso.responsable.id
                        if caso.responsable else None,

                    "FOLDER": caso.folder,
                    "EXP.": caso.expediente,

                    "AGRESOR": caso.agresor,

                    "LATITUD AGRESOR": caso.latitud_agresor,
                    "LONGITUD AGRESOR": caso.longitud_agresor,

                    "TELEFONO": caso.telefono,

                    "fecha_registro":
                        str(caso.fecha_registro)
                        if caso.fecha_registro else "",

                    "ULTIMA VISITA":
                        str(caso.ultima_visita)
                        if caso.ultima_visita else "",

                    "FECHA LIMITE DE SEGUIMIENTO":
                        str(caso.fecha_limite)
                        if caso.fecha_limite else "",

                    "NOTIFICACION BENEFICIARIO":
                        caso.notificacion_beneficiario,

                    "FECHA NOTIFICACION BENEFICIARIO":
                        str(caso.fecha_notificacion_beneficiario)
                        if caso.fecha_notificacion_beneficiario else "",

                    "NOTIFICACION AGRESOR":
                        caso.notificacion_agresor,

                    "FECHA NOTIFICACION AGRESOR":
                        str(caso.fecha_notificacion_agresor)
                        if caso.fecha_notificacion_agresor else "",
                }
            })

        # ==========================================
        # 🟤 UBICACIÓN DEL AGRESOR
        # ==========================================

        if (
            caso.latitud_agresor is not None
            and caso.longitud_agresor is not None
        ):

            features.append({
                "type": "Feature",

                "geometry": {
                    "type": "Point",
                    "coordinates": [
                        caso.longitud_agresor,
                        caso.latitud_agresor
                    ]
                },

                "properties": {
                    "id": caso.id,
                    "TIPO_PUNTO": "AGRESOR",

                    "BENEFICIARIO": caso.beneficiario,
                    "DOMICILIO": caso.domicilio,
                    "NIVEL RIESGO": caso.nivel_riesgo,
                    "DISTRITO": str(caso.distrito_denuncia)
                        if caso.distrito_denuncia else "",

                    "COMISARIA DENUNCIA": caso.comisaria_denuncia,
                    "COMISARIA MEDIDA": caso.comisaria_medida,

                    "EFECTIVO": caso.efectivo,
                    "RESPONSABLE_ID":
                        caso.responsable.id
                        if caso.responsable else None,

                    "FOLDER": caso.folder,
                    "EXP.": caso.expediente,

                    "AGRESOR": caso.agresor,
                    "TELEFONO": caso.telefono,

                    "fecha_registro":
                        str(caso.fecha_registro)
                        if caso.fecha_registro else "",

                    "ULTIMA VISITA":
                        str(caso.ultima_visita)
                        if caso.ultima_visita else "",

                    "FECHA LIMITE DE SEGUIMIENTO":
                        str(caso.fecha_limite)
                        if caso.fecha_limite else "",

                    "NOTIFICACION BENEFICIARIO":
                        caso.notificacion_beneficiario,

                    "FECHA NOTIFICACION BENEFICIARIO":
                        str(caso.fecha_notificacion_beneficiario)
                        if caso.fecha_notificacion_beneficiario else "",

                    "NOTIFICACION AGRESOR":
                        caso.notificacion_agresor,

                    "FECHA NOTIFICACION AGRESOR":
                        str(caso.fecha_notificacion_agresor)
                        if caso.fecha_notificacion_agresor else "",
                }
            })

    return JsonResponse({
        "type": "FeatureCollection",
        "features": features
    })

@login_required
def editar_caso(request, id):

    caso = get_object_or_404(Caso, pk=id)

    # =====================================================
    # DATOS ANTERIORES
    # =====================================================

    responsable_anterior = caso.responsable

    # =====================================================
    # ROLES
    # =====================================================

    es_admin = (
        request.user.is_superuser
        or request.user.groups.filter(
            name="Administrador"
        ).exists()
    )

    es_jefe = request.user.groups.filter(
        name="Jefe_MP"
    ).exists()

    es_usuario = request.user.groups.filter(
        name="Usuario_MP"
    ).exists()

    es_efectivo_comfamea = request.user.groups.filter(
        name="Efectivo_COMFAMEA"
    ).exists()

    es_responsable = (
        caso.responsable == request.user
    )

    # =====================================================
    # EFECTIVO COMFAMEA: NO EDITA
    # =====================================================

    if es_efectivo_comfamea:
        return redirect("gestion_casos")

    # =====================================================
    # ADMINISTRADOR Y JEFE
    # PUEDEN EDITAR SIEMPRE
    # =====================================================

    if es_admin or es_jefe:
        puede_editar = True

    # =====================================================
    # RESPONSABLE
    # =====================================================

    elif es_responsable:

        # -------------------------------------------------
        # AUTORIZACIÓN YA CONCEDIDA
        # -------------------------------------------------

        if caso.edicion_autorizada:
            puede_editar = True

        # -------------------------------------------------
        # PRIMERA EDICIÓN POR NUEVA ASIGNACIÓN
        #
        # IMPORTANTE:
        # SOLO comprobamos el mensaje.
        # NO ponemos edicion_autorizada=True aquí.
        # -------------------------------------------------

        elif Mensaje.objects.filter(
            destinatario=request.user,
            caso=caso,
            asunto="🔔 Nuevo caso asignado",
            leido=False
        ).exists():

            puede_editar = True

        # -------------------------------------------------
        # SIN AUTORIZACIÓN
        # -------------------------------------------------

        else:

            return redirect(
                "solicitar_modificacion",
                id=caso.id
            )

    # =====================================================
    # NO ES RESPONSABLE
    # =====================================================

    else:

        return redirect("gestion_casos")

    # =====================================================
    # FORMULARIO
    # =====================================================

    if request.method == "POST":

        form = CasoForm(
            request.POST,
            instance=caso,
            usuario=request.user
        )

        if form.is_valid():

            # =================================================
            # RESPONSABLE NUEVO
            # =================================================

            caso = form.save(commit=False)

            responsable_nuevo = caso.responsable

            if responsable_nuevo:

                nombre = (
                    f"{responsable_nuevo.first_name} "
                    f"{responsable_nuevo.last_name}"
                ).strip()

                caso.efectivo = (
                    nombre
                    if nombre
                    else responsable_nuevo.username
                )

            # =================================================
            # DETECTAR CAMBIO DE RESPONSABLE
            # =================================================

            responsable_cambio = (
                responsable_anterior != responsable_nuevo
            )

            # =================================================
            # SI CAMBIÓ EL RESPONSABLE
            # =================================================

            if responsable_cambio:

                caso.edicion_autorizada = False

            # =================================================
            # SI EL RESPONSABLE ACTUAL GUARDÓ
            #
            # LA AUTORIZACIÓN SE CONSUME AQUÍ
            # =================================================

            elif es_responsable:

                caso.edicion_autorizada = False

            caso.save()

            # =================================================
            # MARCAR MENSAJES DEL CASO COMO LEÍDOS
            # =================================================

            Mensaje.objects.filter(
                destinatario=request.user,
                caso=caso,
                leido=False
            ).update(
                leido=True,
                fecha_lectura=timezone.now()
            )

            # =================================================
            # MENSAJE AL NUEVO RESPONSABLE
            # =================================================

            if (
                responsable_cambio
                and responsable_nuevo
            ):

                Mensaje.objects.create(

                    destinatario=responsable_nuevo,

                    caso=caso,

                    asunto="🔔 Nuevo caso asignado",

                    contenido=(
                        "Se le comunica que se le ha "
                        "asignado el expediente N.° "
                        f"{caso.expediente or caso.folder} "
                        "para realizar la ejecución y "
                        "seguimiento de las medidas de "
                        "protección."
                    )
                )

            # =================================================
            # SOLICITUD DE MODIFICACIÓN APROBADA
            # =================================================

            if es_usuario:

                solicitud = (
                    SolicitudModificacion.objects.filter(
                        caso=caso,
                        solicitante=request.user,
                        estado="APROBADA"
                    )
                    .order_by(
                        "-fecha_autorizacion"
                    )
                    .first()
                )

                if solicitud:

                    solicitud.estado = "UTILIZADA"

                    solicitud.fecha_utilizacion = (
                        timezone.now()
                    )

                    solicitud.save()

                    caso.edicion_autorizada = False

                    caso.save(
                        update_fields=[
                            "edicion_autorizada"
                        ]
                    )

            return redirect("gestion_casos")

    else:

        form = CasoForm(
            instance=caso,
            usuario=request.user
        )

    response = render(
        request,
        "mapa/nuevo_caso.html",
        {
            "form": form
        }
    )

    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"

    return response

@login_required
@grupo_requerido("Administrador", "Jefe_MP")
def archivar_caso(request, id):

    caso = get_object_or_404(Caso, pk=id)

    caso.estado = "ARCHIVADO"
    caso.save()

    return redirect("gestion_casos")

@login_required
@grupo_requerido("Administrador", "Jefe_MP")
def casos_archivados(request):

    casos = Caso.objects.filter(
        estado="ARCHIVADO"
    ).order_by("-id")

    return render(
        request,
        "mapa/casos_archivados.html",
        {
            "casos": casos
        }
    )

@login_required
@grupo_requerido("Administrador", "Jefe_MP")
def restaurar_caso(request, id):

    caso = get_object_or_404(Caso, pk=id)

    caso.estado = "ACTIVO"
    caso.save()

    return redirect("casos_archivados")

@login_required
@grupo_requerido("Administrador")
def eliminar_caso(request, id):

    caso = get_object_or_404(
        Caso,
        pk=id
    )

    caso.estado = "ELIMINADO"
    caso.fecha_eliminacion = timezone.now()
    caso.eliminado_por = request.user

    caso.save(
        update_fields=[
            "estado",
            "fecha_eliminacion",
            "eliminado_por"
        ]
    )

    return redirect("gestion_casos")


@login_required
@grupo_requerido("Administrador")
def casos_eliminados(request):

    casos = Caso.objects.filter(
        estado="ELIMINADO"
    ).order_by("-fecha_eliminacion")

    return render(
        request,
        "mapa/casos_eliminados.html",
        {
            "casos": casos
        }
    )


@login_required
@grupo_requerido("Administrador")
def restaurar_caso_eliminado(request, id):

    caso = get_object_or_404(
        Caso,
        pk=id,
        estado="ELIMINADO"
    )

    caso.estado = "ACTIVO"
    caso.fecha_eliminacion = None
    caso.eliminado_por = None

    caso.save(
        update_fields=[
            "estado",
            "fecha_eliminacion",
            "eliminado_por"
        ]
    )

    return redirect("casos_eliminados")


@login_required
@grupo_requerido("Administrador")
def eliminar_caso_definitivamente(request, id):

    caso = get_object_or_404(
        Caso,
        pk=id,
        estado="ELIMINADO"
    )

    caso.delete()

    return redirect("casos_eliminados")


# =====================================================
# ELIMINAR AUTOMÁTICAMENTE CASOS DESPUÉS DE 30 DÍAS
# =====================================================

def eliminar_casos_vencidos():

    fecha_limite = timezone.now() - timedelta(days=30)

    casos = Caso.objects.filter(
        estado="ELIMINADO",
        fecha_eliminacion__isnull=False,
        fecha_eliminacion__lte=fecha_limite
    )

    cantidad = casos.count()

    casos.delete()

    return cantidad


@login_required
@grupo_requerido("Jefe_MP", "Usuario_MP")
def solicitar_modificacion(request, id):

    caso = get_object_or_404(Caso, pk=id)

    if request.method == "POST":

        form = SolicitudModificacionForm(request.POST)

        if form.is_valid():

            solicitud = form.save(commit=False)
            solicitud.caso = caso
            solicitud.solicitante = request.user
            solicitud.save()

            return redirect("gestion_casos")

    else:

        form = SolicitudModificacionForm()

    return render(
        request,
        "mapa/solicitar_modificacion.html",
        {
            "form": form,
            "caso": caso
        }
    )

@login_required
@grupo_requerido(
    "Administrador",
    "Jefe_MP",
    "Usuario_MP"
)
def mensajes(request):

    mensajes_nuevos = Mensaje.objects.filter(
        destinatario=request.user,
        leido=False
    ).order_by("-fecha")

    historial_mensajes = Mensaje.objects.filter(
        destinatario=request.user,
        leido=True
    ).order_by("-fecha_lectura", "-fecha")

    solicitudes = SolicitudModificacion.objects.filter(
        estado="PENDIENTE"
    ).order_by("-fecha")


    hoy = timezone.now().date()
    limite = hoy + timedelta(days=3)

    casos_por_vencer = Caso.objects.filter(
        estado="ACTIVO",
        fecha_limite__isnull=False,
        fecha_limite__gte=hoy,
        fecha_limite__lte=limite
    ).filter(
        Q(fecha_registro__year__gte=2026) |
        Q(ultima_visita__year__gte=2026)
    ).order_by("fecha_limite")

    return render(
        request,
        "mapa/mensajes.html",
        {
            "mensajes_nuevos": mensajes_nuevos,
            "historial_mensajes": historial_mensajes,
            "solicitudes": solicitudes,
            "casos_por_vencer": casos_por_vencer,
        }
    )

def cerrar_sesion(request):
    logout(request)
    return redirect("/accounts/login/")

@login_required
def buscar_caso(request):

    texto = request.GET.get("q", "").strip()
    tipo = request.GET.get("tipo", "").strip()

    if texto == "":
        return JsonResponse({
            "encontrado": False,
            "casos": []
        })

    # =====================================================
    # BUSCAR CASOS EXISTENTES
    # =====================================================

    filtros = {
        "beneficiario": Q(beneficiario__icontains=texto),
        "agresor": Q(agresor__icontains=texto),
        "dni": Q(dni_beneficiario__icontains=texto),
        "expediente": Q(expediente__icontains=texto),
        "folder": Q(folder__icontains=texto),
    }

    if tipo in filtros:

        casos = Caso.objects.filter(
            filtros[tipo]
        ).order_by("beneficiario")

    else:

        casos = Caso.objects.filter(
            Q(beneficiario__icontains=texto) |
            Q(agresor__icontains=texto) |
            Q(dni_beneficiario__icontains=texto) |
            Q(dni_agresor__icontains=texto) |
            Q(expediente__icontains=texto) |
            Q(folder__icontains=texto)
        ).order_by("beneficiario")


    resultados = []

    for caso in casos:

        resultados.append({
            "id": caso.id,
            "beneficiario": caso.beneficiario,
            "agresor": caso.agresor,
            "expediente": caso.expediente,
            "folder": caso.folder,
            "latitud": caso.latitud,
            "longitud": caso.longitud,
            "riesgo": caso.nivel_riesgo,
            "tipo": "CASO",
            "url": "/gestion-casos/?q=" + str(caso.beneficiario),
        })


    # =====================================================
    # BUSCAR PERSONAS PRELIMINARES
    # SOLO LAS QUE APARECEN COMO PRESUNTA VÍCTIMA
    # =====================================================

    if tipo in ["", "beneficiario", "dni"]:

        personas_preliminares = PersonaPreliminar.objects.filter(
            convertida=False
        ).filter(
            roles__rol="PRESUNTA_VICTIMA"
        ).distinct()


        if tipo == "beneficiario":

            personas_preliminares = personas_preliminares.filter(
                nombres__icontains=texto
            )

        elif tipo == "dni":

            personas_preliminares = personas_preliminares.filter(
                dni__icontains=texto
            )


        for persona in personas_preliminares:

            if (
                persona.latitud is None
                or persona.longitud is None
            ):
                continue

            resultados.append({

                "id": persona.id,

                "beneficiario":
                    persona.nombres,

                "dni":
                    persona.dni,

                "agresor": "",

                "expediente": "",

                "folder": "",

                "latitud":
                    persona.latitud,

                "longitud":
                    persona.longitud,

                "riesgo":
                    "PRELIMINAR",

                "tipo":
                    "PRELIMINAR",

                "rol":
                    "BENEFICIARIO",

                "url":
                "/casos-preliminares/",

            })


    return JsonResponse({

        "encontrado":
            len(resultados) > 0,

        "casos":
            resultados

    })



@login_required
@grupo_requerido("Administrador", "Jefe_MP")
def aprobar_solicitud(request, id):

    solicitud = get_object_or_404(
        SolicitudModificacion,
        pk=id
    )

    solicitud.estado = "APROBADA"
    solicitud.autorizado_por = request.user
    solicitud.fecha_autorizacion = timezone.now()
    solicitud.save()

    caso = solicitud.caso
    caso.edicion_autorizada = True
    caso.save()

    return redirect("mensajes")

@login_required
@grupo_requerido("Administrador", "Jefe_MP")
def rechazar_solicitud(request, id):

    solicitud = get_object_or_404(
        SolicitudModificacion,
        pk=id
    )

    solicitud.estado = "RECHAZADA"
    solicitud.autorizado_por = request.user
    solicitud.fecha_autorizacion = timezone.now()
    solicitud.save()

    return redirect("mensajes")

@login_required
def mapa_agresores(request):

    agresores = Agresor.objects.filter(
        activo=True,
        latitud__isnull=False,
        longitud__isnull=False
    )

    fecha_limite = timezone.now() - timedelta(days=30)

    agresores_preliminares = PersonaPreliminar.objects.filter(
        convertida=False,
        latitud__isnull=False,
        longitud__isnull=False,
        roles__rol__in=[
            "DENUNCIADO",
            "PARTICIPANTE"
        ]
    ).distinct()

    return render(
        request,
        "mapa/mapa_agresores.html",
        {
            "agresores": agresores,
            "agresores_preliminares": agresores_preliminares,
        }
    )

@login_required
def ubicaciones_preliminares_json(request):

    features = []

    ubicaciones = (
        UbicacionPreliminar.objects
        .filter(
            estado="PRELIMINAR"
        )
        .prefetch_related("personas__roles")
    )

    for ubicacion in ubicaciones:

        personas = ubicacion.personas.filter(
            convertida=False
        ).prefetch_related("roles")

        for persona in personas:

            if persona.convertida:
                continue

            roles = list(
                persona.roles.values_list(
                    "rol",
                    flat=True
                )
            )

            # ==========================================
            # DATOS COMUNES
            # ==========================================

            datos = {
                "id": persona.id,
                "ubicacion_id": ubicacion.id,

                "NOMBRE": persona.nombres or "",
                "DNI": persona.dni or "",
                "DIRECCION": persona.direccion or "",
                "TELEFONO": persona.telefono or "",

                "CONVERTIDA": persona.convertida,

                "ROLES": roles,

                "estado": ubicacion.estado,
                "tipo_registro": ubicacion.tipo_registro,

                "fecha": str(
                    ubicacion.fecha_registro
                ),
            }

            # ==========================================
            # COORDENADAS DE LA PERSONA
            # ==========================================

            if (
                persona.latitud is None
                or persona.longitud is None
            ):
                continue

            lat = persona.latitud
            lng = persona.longitud

            # ==========================================
            # PRESUNTA VÍCTIMA
            # ==========================================

            if (
                "PRESUNTA_VICTIMA" in roles
                and not persona.convertida
            ):

                properties = datos.copy()

                properties.update({

                    "COLOR": "AZUL",

                    "beneficiario":
                        persona.nombres or "",

                    "agresor": "",

                })

                features.append({

                    "type": "Feature",

                    "geometry": {
                        "type": "Point",
                        "coordinates": [
                            lng,
                            lat
                        ]
                    },

                    "properties": properties
                })

            # ==========================================
            # DENUNCIADO
            # ==========================================

            if (
                "DENUNCIADO" in roles
                and not persona.convertida
            ):

                properties = datos.copy()

                properties.update({

                    "COLOR": "MARRON",

                    "beneficiario": "",

                    "agresor":
                        persona.nombres or "",

                })

                features.append({

                    "type": "Feature",

                    "geometry": {
                        "type": "Point",
                        "coordinates": [
                            lng,
                            lat
                        ]
                    },

                    "properties": properties
                })

            # ==========================================
            # PARTICIPANTE
            #
            # UN PARTICIPANTE APARECE EN AMBOS MAPAS
            # ==========================================

            if (
                "PARTICIPANTE" in roles
                and not persona.convertida
            ):

                # ------------------------------
                # 🔵 REPRESENTACIÓN VÍCTIMA
                # ------------------------------

                properties_victima = datos.copy()

                properties_victima.update({

                    "COLOR": "AZUL",

                    "beneficiario":
                        persona.nombres or "",

                    "agresor": "",

                })

                features.append({

                    "type": "Feature",

                    "geometry": {
                        "type": "Point",
                        "coordinates": [
                            lng,
                            lat
                        ]
                    },

                    "properties": properties_victima
                })

                # ------------------------------
                # 🟤 REPRESENTACIÓN AGRESOR
                # ------------------------------

                properties_agresor = datos.copy()

                properties_agresor.update({

                    "COLOR": "MARRON",

                    "beneficiario": "",

                    "agresor":
                        persona.nombres or "",

                })

                features.append({

                    "type": "Feature",

                    "geometry": {
                        "type": "Point",
                        "coordinates": [
                            lng,
                            lat
                        ]
                    },

                    "properties": properties_agresor
                })

    return JsonResponse({

        "type": "FeatureCollection",

        "features": features

    })

@login_required
def estadistica(request):

    # Todos los usuarios pueden visualizar Estadística.
    # La restricción de descarga se controla por separado.
    es_admin = (
        request.user.is_superuser
        or request.user.groups.filter(
            name="Administrador"
        ).exists()
    )

    es_jefe = request.user.groups.filter(
        name="Jefe_MP"
    ).exists()

    es_usuario_estadistica = request.user.groups.filter(
        name="Usuario_Estadistica"
    ).exists()

    puede_descargar = (
        es_admin
        or es_jefe
        or es_usuario_estadistica
    )

    # ==========================================
    # CASOS ACTIVOS
    # ==========================================

    casos_activos = Caso.objects.filter(
        estado="ACTIVO"
    )

    total_casos = casos_activos.count()

    # ==========================================
    # CASOS POR NIVEL DE RIESGO
    # ==========================================

    casos_leves = casos_activos.filter(
        nivel_riesgo="LEVE"
    ).count()

    casos_moderados = casos_activos.filter(
        nivel_riesgo="MODERADO"
    ).count()

    casos_severos = casos_activos.filter(
        nivel_riesgo="SEVERO"
    ).count()

    casos_severo_extremo = casos_activos.filter(
        nivel_riesgo="SEVERO EXTREMO"
    ).count()

    casos_no_determinado = casos_activos.filter(
        nivel_riesgo="NO DETERMINADO"
    ).count()

    # ==========================================
    # CASOS POR VENCER
    # ==========================================

    hoy = timezone.now().date()

    limite = hoy + timedelta(days=3)

    casos_por_vencer = casos_activos.filter(
        fecha_limite__isnull=False,
        fecha_limite__gte=hoy,
        fecha_limite__lte=limite
    ).count()

    # ==========================================
    # CASOS VENCIDOS
    # ==========================================

    casos_vencidos = casos_activos.filter(
        fecha_limite__isnull=False,
        fecha_limite__lt=hoy
    ).count()

    # ==========================================
    # CASOS AL DÍA
    # ==========================================

    casos_al_dia = casos_activos.filter(
        fecha_limite__isnull=False,
        fecha_limite__gt=limite
    ).count()

    # ==========================================
    # PRELIMINARES PENDIENTES
    # ==========================================

    preliminares = UbicacionPreliminar.objects.filter(
        estado="PRELIMINAR"
    )

    total_preliminares = 0

    for ubicacion in preliminares:

        if ubicacion.personas.filter(
            convertida=False
        ).exists():

            total_preliminares += 1

    # ==========================================
    # AGRESORES REGISTRADOS
    # ==========================================

    total_agresores = Agresor.objects.filter(
        activo=True
    ).count()

    # ==========================================
    # CASOS POR EFECTIVO RESPONSABLE
    # ==========================================

    casos_por_efectivo = []

    responsables = User.objects.filter(
        casos_asignados__estado="ACTIVO"
    ).distinct()

    for responsable in responsables:

        casos_efectivo = Caso.objects.filter(
            responsable=responsable,
            estado="ACTIVO"
        )

        casos_por_efectivo.append({
            "id": responsable.id,

            "nombre": responsable.get_full_name() or responsable.username,

            "total": casos_efectivo.count(),

            "leves": casos_efectivo.filter(
                nivel_riesgo="LEVE"
            ).count(),

            "moderados": casos_efectivo.filter(
                nivel_riesgo="MODERADO"
            ).count(),

            "severos": casos_efectivo.filter(
                nivel_riesgo="SEVERO"
            ).count(),

            "severo_extremo": casos_efectivo.filter(
                nivel_riesgo="SEVERO EXTREMO"
            ).count(),

            "no_determinado": casos_efectivo.filter(
                nivel_riesgo="NO DETERMINADO"
            ).count(),

            "por_vencer": casos_efectivo.filter(
                fecha_limite__isnull=False,
                fecha_limite__gte=hoy,
                fecha_limite__lte=limite
            ).count(),

            "vencidos": casos_efectivo.filter(
                fecha_limite__isnull=False,
                fecha_limite__lt=hoy
            ).count(),
        })

    # ==========================================
    # DATOS PARA GRÁFICO DE RIESGO
    # ==========================================

    riesgos = {
        "LEVE": casos_leves,
        "MODERADO": casos_moderados,
        "SEVERO": casos_severos,
        "SEVERO EXTREMO": casos_severo_extremo,
        "NO DETERMINADO": casos_no_determinado,
    }


        # ==========================================
    # COMPARATIVO DE LOS ÚLTIMOS 5 AÑOS
    # ==========================================

    anio_actual = timezone.now().year

    anios_comparativo = list(
        range(anio_actual - 4, anio_actual + 1)
    )

    comparativo_anual = []

    nombres_meses = [
        "",
        "Enero",
        "Febrero",
        "Marzo",
        "Abril",
        "Mayo",
        "Junio",
        "Julio",
        "Agosto",
        "Septiembre",
        "Octubre",
        "Noviembre",
        "Diciembre",
    ]

    for anio in anios_comparativo:

        casos_anio = Caso.objects.filter(
            fecha_registro__year=anio
        )

        meses = []

        for mes in range(1, 13):

            casos_mes = casos_anio.filter(
                fecha_registro__month=mes
            )

            meses.append({
                "mes": mes,

                "nombre": nombres_meses[mes],

                "total": casos_mes.count(),

                "leves": casos_mes.filter(
                    nivel_riesgo="LEVE"
                ).count(),

                "moderados": casos_mes.filter(
                    nivel_riesgo="MODERADO"
                ).count(),

                "severos": casos_mes.filter(
                    nivel_riesgo="SEVERO"
                ).count(),

                "severo_extremo": casos_mes.filter(
                    nivel_riesgo="SEVERO EXTREMO"
                ).count(),

                "no_determinado": casos_mes.filter(
                    nivel_riesgo="NO DETERMINADO"
                ).count(),
            })

        # ==========================================
        # AGREGAR EL AÑO
        # ==========================================

        comparativo_anual.append({

            "anio": anio,

            "total": casos_anio.count(),

            "leves": casos_anio.filter(
                nivel_riesgo="LEVE"
            ).count(),

            "moderados": casos_anio.filter(
                nivel_riesgo="MODERADO"
            ).count(),

            "severos": casos_anio.filter(
                nivel_riesgo="SEVERO"
            ).count(),

            "severo_extremo": casos_anio.filter(
                nivel_riesgo="SEVERO EXTREMO"
            ).count(),

            "no_determinado": casos_anio.filter(
                nivel_riesgo="NO DETERMINADO"
            ).count(),

            "meses": meses,
        })


    return render(
        request,
        "mapa/estadistica.html",
        {
            "total_casos": total_casos,

            "casos_leves": casos_leves,
            "casos_moderados": casos_moderados,
            "casos_severos": casos_severos,
            "casos_severo_extremo": casos_severo_extremo,
            "casos_no_determinado": casos_no_determinado,

            "casos_por_vencer": casos_por_vencer,
            "casos_vencidos": casos_vencidos,
            "casos_al_dia": casos_al_dia,

            "total_preliminares": total_preliminares,

            "total_agresores": total_agresores,

            "casos_por_efectivo": casos_por_efectivo,

            "riesgos": riesgos,

            "puede_descargar": puede_descargar,

            "comparativo_anual": comparativo_anual,
        }
    )

    
@login_required
@grupo_requerido(
    "Administrador",
    "Jefe_MP",
    "Usuario_Estadistica"
)
def casos_por_efectivo(request, id):

    efectivo = get_object_or_404(
        User,
        pk=id
    )

    casos = Caso.objects.filter(
        responsable=efectivo,
        estado="ACTIVO"
    ).order_by("-id")

    return render(
        request,
        "mapa/casos_por_efectivo.html",
        {
            "efectivo": efectivo,
            "casos": casos,
        }
    )

@login_required
def descargar_estadistica(request):

    # ==========================================
    # VERIFICAR AUTORIZACIÓN DE DESCARGA
    # ==========================================

    es_admin = (
        request.user.is_superuser
        or request.user.groups.filter(
            name="Administrador"
        ).exists()
    )

    es_jefe = request.user.groups.filter(
        name="Jefe_MP"
    ).exists()

    es_usuario_estadistica = request.user.groups.filter(
        name="Usuario_Estadistica"
    ).exists()

    if not (
        es_admin
        or es_jefe
        or es_usuario_estadistica
    ):
        return HttpResponseForbidden(
            "No tiene autorización para descargar Estadística."
        )

    # ==========================================
    # IMPORTACIONES
    # ==========================================

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    # ==========================================
    # CASOS ACTIVOS
    # ==========================================

    casos_activos = Caso.objects.filter(
        estado="ACTIVO"
    )

    total_casos = casos_activos.count()

    # ==========================================
    # CASOS POR NIVEL DE RIESGO
    # ==========================================

    casos_leves = casos_activos.filter(
        nivel_riesgo="LEVE"
    ).count()

    casos_moderados = casos_activos.filter(
        nivel_riesgo="MODERADO"
    ).count()

    casos_severos = casos_activos.filter(
        nivel_riesgo="SEVERO"
    ).count()

    casos_severo_extremo = casos_activos.filter(
        nivel_riesgo="SEVERO EXTREMO"
    ).count()

    casos_no_determinado = casos_activos.filter(
        nivel_riesgo="NO DETERMINADO"
    ).count()

    # ==========================================
    # SEGUIMIENTO
    # ==========================================

    hoy = timezone.now().date()

    limite = hoy + timedelta(days=3)

    casos_por_vencer = casos_activos.filter(
        fecha_limite__isnull=False,
        fecha_limite__gte=hoy,
        fecha_limite__lte=limite
    ).count()

    casos_vencidos = casos_activos.filter(
        fecha_limite__isnull=False,
        fecha_limite__lt=hoy
    ).count()

    casos_al_dia = casos_activos.filter(
        fecha_limite__isnull=False,
        fecha_limite__gt=limite
    ).count()

    # ==========================================
    # PRELIMINARES
    # ==========================================

    preliminares = UbicacionPreliminar.objects.filter(
        estado="PRELIMINAR"
    )

    total_preliminares = 0

    for ubicacion in preliminares:

        if ubicacion.personas.filter(
            convertida=False
        ).exists():

            total_preliminares += 1

    # ==========================================
    # AGRESORES
    # ==========================================

    total_agresores = Agresor.objects.filter(
        activo=True
    ).count()

    # ==========================================
    # CREAR EXCEL
    # ==========================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Estadística"

    # ==========================================
    # TÍTULO
    # ==========================================

    ws["A1"] = "ESTADÍSTICA - COMFAM PNP EL AGUSTINO"

    ws["A1"].font = Font(
        bold=True,
        size=14
    )

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )

    ws.merge_cells("A1:B1")

    # ==========================================
    # ENCABEZADOS
    # ==========================================

    ws["A3"] = "INDICADOR"
    ws["B3"] = "CANTIDAD"

    ws["A3"].font = Font(bold=True)
    ws["B3"].font = Font(bold=True)

    # ==========================================
    # DATOS
    # ==========================================

    datos = [

        (
            "Total de casos activos",
            total_casos
        ),

        (
            "Casos leves",
            casos_leves
        ),

        (
            "Casos moderados",
            casos_moderados
        ),

        (
            "Casos severos",
            casos_severos
        ),

        (
            "Casos severo extremo",
            casos_severo_extremo
        ),

        (
            "Casos no determinado",
            casos_no_determinado
        ),

        (
            "Casos por vencer en 3 días",
            casos_por_vencer
        ),

        (
            "Casos vencidos",
            casos_vencidos
        ),

        (
            "Casos al día",
            casos_al_dia
        ),

        (
            "Preliminares pendientes",
            total_preliminares
        ),

        (
            "Agresores registrados",
            total_agresores
        ),

    ]

    fila = 4

    for indicador, cantidad in datos:

        ws.cell(
            row=fila,
            column=1,
            value=indicador
        )

        ws.cell(
            row=fila,
            column=2,
            value=cantidad
        )

        fila += 1

    # ==========================================
    # ANCHO DE COLUMNAS
    # ==========================================

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15

    # ==========================================
    # GENERAR ARCHIVO
    # ==========================================

    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    # ==========================================
    # RESPUESTA HTTP
    # ==========================================

    response = HttpResponse(
        archivo.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="Estadistica_COMFAM.xlsx"'
    )

    return response

@login_required
def corregir_niveles_riesgo(request):

    if not (
        request.user.is_superuser
        or request.user.groups.filter(
            name="Administrador"
        ).exists()
    ):
        return HttpResponseForbidden("No autorizado.")

    from django.db import transaction

    with transaction.atomic():

        Caso.objects.filter(
            nivel_riesgo="SEVERO1"
        ).update(
            nivel_riesgo="SEVERO"
        )

        Caso.objects.filter(
            nivel_riesgo="SEVERO2"
        ).update(
            nivel_riesgo="SEVERO EXTREMO"
        )

    return redirect("estadistica")